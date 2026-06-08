"""
Export matching results to CSV or Excel.

Excel workbook tabs:
  All Matches | Very High Confidence | High Confidence | Medium Confidence |
  Low Confidence | Address Clusters | Match Settings | Extraction Notes
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

_CONF_FILL = {
    'Very High': PatternFill('solid', fgColor='C6EFCE'),
    'High':      PatternFill('solid', fgColor='9DC3E6'),
    'Medium':    PatternFill('solid', fgColor='FFEB9C'),
    'Low':       PatternFill('solid', fgColor='FCE4D6'),
    'Ignore':    PatternFill('solid', fgColor='D9D9D9'),
}

_HEADER_FILL = PatternFill('solid', fgColor='2F5496')
_HEADER_FONT = Font(color='FFFFFF', bold=True)
_REVIEW_FILL = PatternFill('solid', fgColor='FFD7D7')
_THIN = Side(style='thin', color='AAAAAA')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

DISCLAIMER = (
    "INVESTIGATIVE LEADS ONLY — These results identify potential address matches "
    "for investigative purposes. They do not constitute proof of chameleon-carrier "
    "behavior, insurance fraud, or any other misconduct. "
    "All matches require manual verification before any investigative or "
    "enforcement action is taken."
)


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode('utf-8')


def _write_sheet(ws, df: pd.DataFrame, confidence_col: str = 'confidence',
                 review_col: str = 'review_needed') -> None:
    """Write a DataFrame to a worksheet with header styling and row coloring."""
    if df.empty:
        ws.append(['No data'])
        return

    # Header row
    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for _, row in df.iterrows():
        ws.append(['' if (isinstance(v, float) and v != v) else v
                   for v in row])
        excel_row = ws.max_row
        conf = str(row.get(confidence_col, ''))
        needs_review = str(row.get(review_col, '')).lower() in ('true', '1', 'yes')
        fill = _CONF_FILL.get(conf)

        for cell in ws[excel_row]:
            if fill:
                cell.fill = fill
            if needs_review and not fill:
                cell.fill = _REVIEW_FILL
            cell.alignment = Alignment(wrap_text=False)

    # Auto-fit column widths (capped at 60)
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1), 1):
        max_len = max((len(str(c.value or '')) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def to_excel_bytes(
    matches_df: pd.DataFrame,
    clusters_df: pd.DataFrame | None = None,
    match_settings: dict | None = None,
    extraction_notes: str = '',
) -> bytes:
    output = io.BytesIO()
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Tab order
    conf_levels = ['Very High', 'High', 'Medium', 'Low']

    # --- All Matches ---
    ws_all = wb.create_sheet('All Matches')
    _write_sheet(ws_all, matches_df)

    # --- Confidence tabs ---
    for level in conf_levels:
        subset = (
            matches_df[matches_df['confidence'] == level].copy()
            if not matches_df.empty and 'confidence' in matches_df.columns
            else pd.DataFrame()
        )
        ws = wb.create_sheet(f'{level} Confidence')
        _write_sheet(ws, subset)

    # --- Address Clusters ---
    ws_clust = wb.create_sheet('Address Clusters')
    clust = clusters_df if (clusters_df is not None and not clusters_df.empty) else pd.DataFrame()
    _write_sheet(ws_clust, clust, confidence_col='', review_col='review_flag')

    # --- Match Settings ---
    ws_set = wb.create_sheet('Match Settings')
    settings_rows = [['Setting', 'Value']]
    for cell in ws_set[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    if match_settings:
        for k, v in match_settings.items():
            settings_rows.append([k, str(v)])
    for row in settings_rows:
        ws_set.append(row)
    ws_set.column_dimensions['A'].width = 30
    ws_set.column_dimensions['B'].width = 50

    # --- Extraction Notes ---
    ws_notes = wb.create_sheet('Extraction Notes')
    ws_notes.append(['Category', 'Note'])
    for cell in ws_notes[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws_notes.append(['DISCLAIMER', DISCLAIMER])
    ws_notes.append(['Extraction Notes', extraction_notes or 'N/A'])
    ws_notes.column_dimensions['A'].width = 20
    ws_notes.column_dimensions['B'].width = 100
    for row in ws_notes.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(output)
    output.seek(0)
    return output.read()
